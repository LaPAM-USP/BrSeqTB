#!/usr/bin/env python3
# ============================================================
# Generate manifest.tsv from input_table.csv or input_table.xlsx
# and validate reads/
#
# Default input:
# input/input_table.csv
#
# CSV rules:
# - comma-separated
# - quotes are optional
# - quotes are only needed if a field contains comma, quote or newline
#
# Required column:
# Biosample
#
# Expected Illumina naming convention:
# <BIOSAMPLE>_S<NUM>_L<NNN>_R1_001.fastq.gz
# <BIOSAMPLE>_S<NUM>_L<NNN>_R2_001.fastq.gz
# ============================================================

import argparse
import csv
import os
import re
import sys
from collections import Counter, defaultdict
from typing import Optional, List, Dict, Tuple


def die(msg: str, code: int = 1) -> None:
    print(f"[ERROR] {msg}", file=sys.stderr)
    sys.exit(code)


def read_text_file_robust(table_path: str) -> Tuple[str, str]:
    encodings = ["utf-8-sig", "utf-16", "cp1252", "latin1"]

    for enc in encodings:
        try:
            with open(table_path, "r", encoding=enc, newline="") as f:
                return f.read(), enc
        except UnicodeDecodeError:
            continue

    die(f"Could not decode input table: {table_path}. Please save it as UTF-8.")


def read_biosamples_from_csv(csv_path: str) -> List[str]:
    if not os.path.exists(csv_path):
        die(f"input_table.csv not found: {csv_path}")

    content, encoding_used = read_text_file_robust(csv_path)

    if not content.strip():
        die("input_table.csv is empty.")

    first_line = content.splitlines()[0]

    if "\t" in first_line and "," not in first_line:
        die(
            "Input table appears to be TAB-separated, but CSV comma-separated is required. "
            "Please export as CSV using comma as separator."
        )

    if ";" in first_line and "," not in first_line:
        die(
            "Input table appears to be semicolon-separated (;), but CSV comma-separated is required. "
            "Please export as CSV using comma as separator."
        )

    reader = csv.DictReader(content.splitlines())

    if reader.fieldnames is None:
        die("input_table.csv has no header.")

    reader.fieldnames = [
        h.strip().lstrip("\ufeff").strip('"\'') if h is not None else h
        for h in reader.fieldnames
    ]

    if "Biosample" not in reader.fieldnames:
        die("Required column 'Biosample' not found in input_table.csv.")

    biosamples: List[str] = []

    for i, row in enumerate(reader, start=2):
        s = (row.get("Biosample") or "").strip().strip('"\'')

        if not s:
            die(f"Missing required Biosample value in input_table.csv at line {i}.")

        biosamples.append(s)

    if not biosamples:
        die("No biosamples found in the 'Biosample' column of input_table.csv.")

    c = Counter(biosamples)
    dups = [b for b, n in c.items() if n > 1]
    if dups:
        die(f"Duplicate biosample IDs found in input_table.csv: {', '.join(dups)}")

    print(f"[OK] input_table.csv loaded using encoding: {encoding_used}")

    return biosamples


def read_biosamples_from_xlsx(xlsx_path: str, sheet: Optional[str] = None) -> List[str]:
    if not os.path.exists(xlsx_path):
        die(f"input_table.xlsx not found: {xlsx_path}")

    try:
        from openpyxl import load_workbook
    except ImportError:
        die("Missing dependency: openpyxl. Install it only if you want to use .xlsx input.", code=2)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]

    biosamples: List[str] = []

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue

        if all(v is None or str(v).strip() == "" for v in row):
            continue

        v = row[0]

        if v is None:
            if i == 1:
                continue
            die(f"Missing required Biosample value in input_table.xlsx at row {i}.")

        s = str(v).strip().strip('"\'')

        if not s:
            if i == 1:
                continue
            die(f"Missing required Biosample value in input_table.xlsx at row {i}.")

        if i == 1 and s.lower() == "biosample":
            continue

        biosamples.append(s)

    wb.close()

    if not biosamples:
        die("No biosamples found in the first column of input_table.xlsx after header.")

    c = Counter(biosamples)
    dups = [b for b, n in c.items() if n > 1]
    if dups:
        die(f"Duplicate biosample IDs found in input_table.xlsx: {', '.join(dups)}")

    return biosamples


def read_biosamples_from_table(table_path: str, sheet: Optional[str] = None) -> List[str]:
    if table_path.endswith(".csv"):
        return read_biosamples_from_csv(table_path)

    if table_path.endswith(".xlsx"):
        return read_biosamples_from_xlsx(table_path, sheet)

    die("Input table must have extension .csv or .xlsx")


def validate_reads_collect_errors(
    reads_dir: str,
    biosamples: List[str]
) -> Tuple[Dict[str, Dict[Tuple[str, str], Dict[str, str]]], List[str]]:

    errors: List[str] = []

    if not os.path.isdir(reads_dir):
        return {}, [f"[FATAL] reads directory not found: {reads_dir}"]

    try:
        all_files = os.listdir(reads_dir)
    except Exception as e:
        return {}, [f"[FATAL] Cannot list reads directory '{reads_dir}': {e}"]

    all_fastqs = [f for f in all_files if f.endswith(".fastq.gz")]

    results: Dict[str, Dict[Tuple[str, str], Dict[str, str]]] = {}

    for biosample in biosamples:
        pattern = re.compile(
            r"^" + re.escape(biosample) + r"_S(\d+)_L(\d{3})_R([12])_001\.fastq\.gz$"
        )

        candidates = [f for f in all_fastqs if f.startswith(biosample + "_")]

        if not candidates:
            errors.append(
                f"Biosample '{biosample}': no FASTQs in '{reads_dir}'. "
                f"Expected e.g.: {biosample}_S1_L001_R1_001.fastq.gz and R2."
            )
            continue

        valid: List[str] = []
        invalid: List[str] = []
        for f in candidates:
            if pattern.match(f):
                valid.append(f)
            else:
                invalid.append(f)

        if invalid:
            errors.append(
                "Biosample '{b}': invalid naming (Illumina convention required):\n{lst}".format(
                    b=biosample,
                    lst="\n".join([f"  - {x}" for x in sorted(invalid)])
                )
            )

        if not valid:
            continue

        pairs: Dict[Tuple[str, str], Dict[str, str]] = defaultdict(dict)
        for f in valid:
            m = pattern.match(f)
            if not m:
                continue
            s_num, lane, read = m.group(1), m.group(2), m.group(3)
            key = (s_num, lane)
            pairs[key][f"R{read}"] = os.path.join(reads_dir, f)

        missing: List[str] = []
        for (s_num, lane), rr in sorted(pairs.items()):
            if "R1" not in rr:
                missing.append(
                    f"missing R1 for S{s_num} L{lane} (expected: {biosample}_S{s_num}_L{lane}_R1_001.fastq.gz)"
                )
            if "R2" not in rr:
                missing.append(
                    f"missing R2 for S{s_num} L{lane} (expected: {biosample}_S{s_num}_L{lane}_R2_001.fastq.gz)"
                )

        if missing:
            errors.append(
                "Biosample '{b}': unpaired FASTQs:\n{lst}".format(
                    b=biosample,
                    lst="\n".join([f"  - {x}" for x in missing])
                )
            )
            continue

        results[biosample] = dict(pairs)

    return results, errors


def write_manifest(out_path: str, biosamples: List[str]) -> None:
    out_dir = os.path.dirname(out_path) or "."
    os.makedirs(out_dir, exist_ok=True)

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("biosample\n")
        for b in biosamples:
            f.write(f"{b}\n")


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Generate manifest.tsv from input_table.csv or input_table.xlsx and validate reads/."
    )
    ap.add_argument("--table", default="input/input_table.csv", help="Path to input_table.csv or input_table.xlsx")
    ap.add_argument("--sheet", default=None, help="Excel sheet name, only used for .xlsx input")
    ap.add_argument("--reads", default="reads", help="Reads directory (default: reads)")
    ap.add_argument(
        "--out",
        default="manifest.tsv",
        help="Output manifest TSV (default: manifest.tsv)"
    )
    args = ap.parse_args()

    biosamples = read_biosamples_from_table(args.table, args.sheet)

    pairs, errs = validate_reads_collect_errors(args.reads, biosamples)

    if errs:
        print("[ERROR] Validation failed. Fix the following issues and re-run:\n", file=sys.stderr)
        for e in errs:
            print(f"- {e}\n", file=sys.stderr)
        print(f"[ERROR] Total biosamples with issues: {len(errs)} / {len(biosamples)}", file=sys.stderr)
        return 1

    write_manifest(args.out, biosamples)

    total_pairs = sum(len(pairs[b]) for b in biosamples)
    total_fastq = total_pairs * 2

    print("[OK] Manifest generated and reads validated.")
    print(f"     Biosamples: {len(biosamples)}")
    print(f"     Total (S,L) pairs: {total_pairs}")
    print(f"     Total FASTQs matched: {total_fastq}")
    print(f"     Manifest: {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())